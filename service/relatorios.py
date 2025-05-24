from model.models import Protesto, Empresa
from typing import Dict, Any, List
import re
from model.models import ReplanilhamentoHistorico, User
from sqlalchemy.orm import Session
import shutil
import openpyxl
from datetime import datetime
from calendar import month_abbr
from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks, Body, Query
from fastapi.responses import FileResponse
import os
import httpx

def preencher_faturamento(dados_faturamento, caminho_saida):

   # Abre a planilha copiada
   wb = openpyxl.load_workbook(caminho_saida)
   ws = wb["FATURAMENTO"]  # Garante aba correta

   # Mapeamento de meses e colunas
   meses = {
      "jan": 2, "fev": 3, "mar": 4, "abr": 5, "mai": 6, "jun": 7,
      "jul": 8, "ago": 9, "set": 10, "out": 11, "nov": 12, "dez": 13
   }
   anos = {
      "2022": 3,
      "2023": 4,
      "2024": 5,
      "2025": 6,
   }

   # Preenche a planilha
   for chave, valor in dados_faturamento.items():
      try:
         mes_abrev, ano = chave.split("/")
         mes_abrev = mes_abrev.lower()
         if mes_abrev in meses and ano in anos:
            row = meses[mes_abrev] + 1
            col = anos[ano]
            ws.cell(row=row, column=col, value=valor)
      except Exception as e:
         print(f"Erro ao preencher '{chave}': {e}")

   # Salva planilha preenchida
   wb.save(caminho_saida)


def preencher_endividamento(dados_por_mes: Dict[str, list], caminho_saida: str):
   wb = openpyxl.load_workbook(caminho_saida)
   ws = wb["DÍVIDA"]

   # Mapeia nome do mês para número
   mes_nome_para_numero = {name.lower(): idx for idx, name in enumerate(month_abbr) if name}

   def mes_para_ordem(mes: str):
      try:
         parte_mes, parte_ano = mes.lower().split("/")
         mes_num = mes_nome_para_numero.get(parte_mes[:3], 0)
         return (int(parte_ano), mes_num)
      except Exception:
         return (9999, 99)

   # Ordena meses
   meses = sorted(dados_por_mes.keys(), key=mes_para_ordem)

   # === 🟦 TABELA 1: Operações individuais, colunas por mês ===
   ws.cell(row=1, column=1, value="Banco")
   ws.cell(row=1, column=2, value="Tipo")
   for idx, mes in enumerate(meses):
      ws.cell(row=1, column=3 + idx, value=mes)

   linha = 2
   for mes_idx, mes in enumerate(meses):
      col = 3 + mes_idx
      for item in dados_por_mes[mes]:
         banco = item.get("fundo")
         tipo = item.get("tipo")
         valor = item.get("valor", 0.0)

         ws.cell(row=linha, column=1, value=banco)
         ws.cell(row=linha, column=2, value=tipo)
         ws.cell(row=linha, column=col, value=valor)
         linha += 1

   # 🟩 SOMA FINAL da Tabela 1
   ws.cell(row=linha, column=2, value="TOTAL")
   for idx in range(len(meses)):
      col = 3 + idx
      cell_range = f"{openpyxl.utils.get_column_letter(col)}2:{openpyxl.utils.get_column_letter(col)}{linha - 1}"
      ws.cell(row=linha, column=col, value=f"=SUM({cell_range})")

   # === 🟨 TABELA 2: Agrupado por banco + tipo ===
   col_offset = 2 + len(meses) + 3
   ws.cell(row=1, column=col_offset, value="Banco")
   ws.cell(row=1, column=col_offset + 1, value="Tipo")
   for idx, mes in enumerate(meses):
      ws.cell(row=1, column=col_offset + 2 + idx, value=mes)

   totais_banco_tipo = {}
   for mes in meses:
      for item in dados_por_mes[mes]:
         banco = item.get("fundo")
         tipo = item.get("tipo")
         valor = item.get("valor", 0.0)
         chave = (banco, tipo)

         if chave not in totais_banco_tipo:
            totais_banco_tipo[chave] = {}
         if mes not in totais_banco_tipo[chave]:
            totais_banco_tipo[chave][mes] = 0
         totais_banco_tipo[chave][mes] += valor

   linha_resumo = 2
   for (banco, tipo), valores_por_mes in totais_banco_tipo.items():
      ws.cell(row=linha_resumo, column=col_offset, value=banco)
      ws.cell(row=linha_resumo, column=col_offset + 1, value=tipo)
      for idx, mes in enumerate(meses):
         valor = valores_por_mes.get(mes, 0)
         ws.cell(row=linha_resumo, column=col_offset + 2 + idx, value=valor)
      linha_resumo += 1

   # 🟩 SOMA FINAL da Tabela 2
   ws.cell(row=linha_resumo, column=col_offset + 1, value="TOTAL")
   for idx in range(len(meses)):
      col = col_offset + 2 + idx
      cell_range = f"{openpyxl.utils.get_column_letter(col)}2:{openpyxl.utils.get_column_letter(col)}{linha_resumo - 1}"
      ws.cell(row=linha_resumo, column=col, value=f"=SUM({cell_range})")

   wb.save(caminho_saida)
   wb.close()


def extrair_arquivos_com_request_id(data: list) -> list:
   arquivos_resultado = []

   for item in data:
      request_id = item.get("request_id")
      arquivos = item.get("result", {}).get("arquivos", [])
      for arquivo in arquivos:
         arquivo_com_id = dict(arquivo)  # copia o dicionário
         arquivo_com_id["request_id"] = request_id
         arquivos_resultado.append(arquivo_com_id)

   return arquivos_resultado


def filtrar_arquivos_por_request_id(dados, user_request_ids):
   resultado_filtrado = []

   for item in dados:
      request_id = item.get("request_id")

      # Se o request_id for de um request permitido pelo usuário
      if request_id in user_request_ids:
         # Mantém o item, pois pertence ao usuário
         resultado_filtrado.append(item)
      else:
         # Caso queira manter o item mas remover só os arquivos (não parece ser seu caso)
         # item["result"]["arquivos"] = []
         # resultado_filtrado.append(item)
         pass  # remove completamente o item

   return resultado_filtrado

async def gerar_relatorio_db(
        tipo_relatorio: str,
        mes_inicio: str,
        mes_fim: str,
        cnpj: str,
        user_id,
        db: Session
):
   user = db.query(User).filter(User.id == user_id).first()

   user_org_id = user.organizacao_id

   result = db.query(ReplanilhamentoHistorico.ms_replanning_request_id) \
      .filter(
      ReplanilhamentoHistorico.cnpj == cnpj,
      ReplanilhamentoHistorico.organizacao_id == user_org_id,
   ) \
      .order_by(ReplanilhamentoHistorico.data_upload.desc()) \
      .all()

   user_request_ids = [r[0] for r in result]

   arquivos_validados_faturamento = await buscar_arquivos_do_replanning('faturamento', mes_inicio,mes_fim)
   #arquivos_validados_faturamento = filtrar_arquivos_por_organizacao(mocked_faturamento.get("arquivos", []),user_org_id,cnpj, db)

   arquivos_validados_endividamento = await buscar_arquivos_do_replanning('endividamento', mes_inicio,mes_fim)
   #arquivos_validados_endividamento = filtrar_arquivos_por_organizacao(mocked_endividamento.get("arquivos", []),user_org_id,cnpj, db)

   arquivos_do_usuario_faturamento = filtrar_arquivos_por_request_id(arquivos_validados_faturamento, user_request_ids)
   arquivos_do_usuario_endividamento = filtrar_arquivos_por_request_id(arquivos_validados_endividamento, user_request_ids)

   dados_faturamento = organizar_faturamento_mes_a_mes({"arquivos": extrair_arquivos_com_request_id(arquivos_do_usuario_faturamento)})
   dados_endividamento = organizar_endividamento_mes_a_mes({"arquivos": extrair_arquivos_com_request_id(arquivos_do_usuario_endividamento)})

   caminho_modelo = "utils/modelo_larca_planilhamento.xlsx"

   now = datetime.now()
   timestamp = now.strftime("%Y-%m-%d_%H-%M-%S") + f"-{int(now.microsecond / 1000):03d}"
   caminho_saida = f"utils/planilha_preenchida_{timestamp}.xlsx"

   shutil.copyfile(caminho_modelo, caminho_saida)

   preencher_faturamento(dados_faturamento,caminho_saida)
   preencher_endividamento(dados_endividamento, caminho_saida)

   # 5. Retorna o arquivo como resposta e agenda remoção
   response = FileResponse(
      caminho_saida,
      filename=f"relatorio_{tipo_relatorio}_{timestamp}.xlsx",
      media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
      headers={"Cache-Control": "no-cache"}
   )

   # Envolve o FileResponse para remover o arquivo após envio
   def remover_arquivo_após_envio(path: str):
      try:
         os.remove(path)
      except Exception as e:
         print(f"[WARN] Falha ao remover arquivo temporário: {e}")

   # Agendando a remoção após envio (via background task se FastAPI usar)
   import threading
   threading.Timer(5.0, remover_arquivo_após_envio, args=[caminho_saida]).start()

   return response

##UTILS
async def buscar_arquivos_do_replanning(
    process_type: str,
    mes_inicio: str,
    mes_fim: str
) -> list:
    url = "https://replanning-api-fc1332dd48fb.herokuapp.com/process-results"
    params = {
        "process_type": process_type,
        "start_date": mes_inicio,
        "end_date": mes_fim
    }
    print(params)

    async with httpx.AsyncClient(follow_redirects=True) as client:
        response = await client.get(url, params=params)
        response.raise_for_status()
        return response.json()

def filtrar_arquivos_por_organizacao(
        arquivos: List[Dict],
        user_org_id: int,
        cnpj:str,
        db: Session
) -> List[Dict]:
   arquivos_validados = []

   for arquivo in arquivos:
      request_id = arquivo.get("request_id")
      if not request_id:
         continue  # ignora arquivos sem request_id

      historico = (
         db.query(ReplanilhamentoHistorico)
         .filter(ReplanilhamentoHistorico.ms_replanning_request_id == request_id)
         .filter(ReplanilhamentoHistorico.cnpj == cnpj)
         .first()
      )

      if historico and historico.organizacao_id == user_org_id and historico.cnpj == cnpj:
         arquivos_validados.append(arquivo)

   return arquivos_validados

def organizar_faturamento_mes_a_mes(data: Dict) -> Dict[str, float]:
   faturamento_por_mes = {}

   for arquivo in data.get("arquivos", []):
      for entrada in arquivo.get("faturamento_mensal", []):
         mes_str = entrada.get("mes", "").strip().lower()

         # Normalização: substitui 'maio' por 'mai', 'setembro' por 'set', etc.
         mes_str = mes_str.replace("maio", "mai")
         mes_str = mes_str.replace("setembro", "set")

         # Usa regex para extrair mês abreviado e ano
         match = re.match(r"([a-zç]{3})/?(\d{4})", mes_str)
         if match:
            mes_abrev, ano = match.groups()
            chave = f"{mes_abrev}/{ano}"
            if chave not in faturamento_por_mes:
               faturamento_por_mes[chave] = entrada["valor"]

   # Retorna os meses ordenados por ano e mês
   def ordenar_key(k):
      mes_map = {
         "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5,
         "jun": 6, "jul": 7, "ago": 8, "set": 9,
         "out": 10, "nov": 11, "dez": 12
      }
      mes, ano = k.split("/")
      return (int(ano), mes_map.get(mes[:3], 0))

   faturamento_ordenado = dict(sorted(faturamento_por_mes.items(), key=lambda item: ordenar_key(item[0])))

   return faturamento_ordenado


def organizar_endividamento_mes_a_mes(dados: Dict[str, Any]) -> Dict[str, List[Dict[str, Any]]]:
   """
   Organiza os dados de endividamento por mês, preservando cada entrada como ela aparece no arquivo.
   Não agrupa por fundo ou tipo.

   :param dados: JSON com arquivos contendo endividamento
   :return: dicionário com chave = mês, e valor = lista de itens de endividamento
   """
   resultado = {}
   meses_processados = set()

   for arquivo in dados.get("arquivos", []):
      mes = arquivo.get("data")

      if mes in meses_processados:
         continue  # pula meses repetidos

      meses_processados.add(mes)
      resultado[mes] = []

      for item in arquivo.get("endividamento", []):
         resultado[mes].append({
            "fundo": item.get("fundo"),
            "tipo": item.get("tipo"),
            "valor": item.get("valor", 0.0)
         })

   return resultado

mocked_faturamento = {
   "empresa":"AGROPARR ALIMENTOS LTDA.",
   "cnpj":"93.607.398/0001-00",
   "regime_tributario":"LUCRO REAL",
   "arquivos":[
      {
         "nome":"Agroparr -  Faturamento  2025.pdf",
         "periodo":"jan/2025 a mar/2025",
         "localidade":"Sentinela do Sul/RS",
         "faturamento_mensal":[
            {
               "mes":"jan/2025",
               "valor":5646344.78
            },
            {
               "mes":"fev/2025",
               "valor":6424275.23
            },
            {
               "mes":"mar/2025",
               "valor":19093930.72
            }
         ],
         "total_faturado":31295050.73,
         "request_id": "fa70b143-07e6-477d-bfd7-c6b2e18ea33a"
      },
      {
         "nome":"FATURAMENTO ULTIMOS 12 MESES.pdf",
         "periodo":"jan/2024 a dez/2024",
         "localidade":"Ven\u00e2ncio Aires \u2013 RS",
         "faturamento_mensal":[
            {
               "mes":"jan/2024",
               "valor":16548706.71
            },
            {
               "mes":"fev/2024",
               "valor":14763816.72
            },
            {
               "mes":"mar/2024",
               "valor":14385627.98
            },
            {
               "mes":"abr/2024",
               "valor":14817490.45
            },
            {
               "mes":"mai/2024",
               "valor":10708817.61
            },
            {
               "mes":"jun/2024",
               "valor":10795534.85
            },
            {
               "mes":"jul/2024",
               "valor":12935328.67
            },
            {
               "mes":"ago/2024",
               "valor":11934619.35
            },
            {
               "mes":"set/2024",
               "valor":9777860.16
            },
            {
               "mes":"out/2024",
               "valor":9351795.79
            },
            {
               "mes":"nov/2024",
               "valor":7936779.16
            },
            {
               "mes":"dez/2024",
               "valor":6059322.02
            }
         ],
         "total_faturado":140015699.47,
         "request_id": "fa70b143-07e6-477d-bfd7-c6b2e18ea33a"
      },
      {
         "nome":"Condusvale - Faturamento 2023 a 08.2024.pdf",
         "periodo":"jan/2023 a dez/2024",
         "localidade":"Ven\u00e2ncio Aires \u2013 RS",
         "faturamento_mensal":[
            {
               "mes":"jan/2023",
               "valor":18264431.92
            },
            {
               "mes":"fev/2023",
               "valor":14210470.87
            },
            {
               "mes":"mar/2023",
               "valor":17413088.09
            },
            {
               "mes":"abr/2023",
               "valor":13225801.75
            },
            {
               "mes":"maio/2023",
               "valor":15466675.86
            },
            {
               "mes":"jun/2023",
               "valor":14622745.85
            },
            {
               "mes":"jul/2023",
               "valor":15951969.88
            },
            {
               "mes":"ago/2023",
               "valor":17717871.24
            },
            {
               "mes":"set/2023",
               "valor":14324459.55
            },
            {
               "mes":"out/2023",
               "valor":17024119.83
            },
            {
               "mes":"nov/2023",
               "valor":19555487.58
            },
            {
               "mes":"dez/2023",
               "valor":15085978.5
            },
            {
               "mes":"jan/2024",
               "valor":16548706.71
            },
            {
               "mes":"fev/2024",
               "valor":14763816.72
            },
            {
               "mes":"mar/2024",
               "valor":14385627.98
            },
            {
               "mes":"abr/2024",
               "valor":14817490.45
            },
            {
               "mes":"maio/2024",
               "valor":10708817.61
            },
            {
               "mes":"jun/2024",
               "valor":10795534.85
            },
            {
               "mes":"jul/2024",
               "valor":12935328.67
            },
            {
               "mes":"ago/2024",
               "valor":11934619.35
            }
         ],
         "total_faturado":299753043.26,
         "request_id": "e3073a12-0136-454f-b87e-44ec4acdf60b"
      }
   ]
}


mocked_endividamento = {
   "arquivos":[
      {
         "nome":"Dacalda - Endividamento 11.2024.pdf",
         "data":"nov/2024",
         "endividamento":[
            {
               "fundo":"Opea Securitiza\u00e7\u00e3o",
               "tipo":"CRA",
               "valor":80408784.0
            },
            {
               "fundo":"BANCO BBM VJ",
               "tipo":"CCB - Linha FGI",
               "valor":9343123.0
            },
            {
               "fundo":"BANCO BBM VJ",
               "tipo":"CCB - Linha FGI",
               "valor":9879803.0
            },
            {
               "fundo":"BANCO BBM",
               "tipo":"CCB - Linha FGI",
               "valor":9343123.0
            },
            {
               "fundo":"BANCO BBM",
               "tipo":"CCB - Linha FGI",
               "valor":11546644.0
            },
            {
               "fundo":"BANCO BBM",
               "tipo":"CCB - Linha FGI",
               "valor":8657841.0
            },
            {
               "fundo":"BANCO BRADESCO",
               "tipo":"CCB",
               "valor":5467618.0
            },
            {
               "fundo":"BANCO BRADESCO",
               "tipo":"CPR",
               "valor":3240697.0
            },
            {
               "fundo":"BANCO DO BRASIL",
               "tipo":"CDCA",
               "valor":27522621.0
            },
            {
               "fundo":"BANCO DO BRASIL",
               "tipo":"CDCA",
               "valor":4006740.0
            },
            {
               "fundo":"BANCO DO BRASIL",
               "tipo":"CDCA",
               "valor":6278027.0
            },
            {
               "fundo":"BANCO DO BRASIL",
               "tipo":"CCB",
               "valor":4869721.0
            },
            {
               "fundo":"BANCO FIBRA",
               "tipo":"CCB - Linha FGI",
               "valor":2672452.0
            },
            {
               "fundo":"BANCO ITA\u00da",
               "tipo":"CPR",
               "valor":7582066.0
            },
            {
               "fundo":"BANCO ITA\u00da",
               "tipo":"CPR",
               "valor":2316891.0
            },
            {
               "fundo":"BANCO ITA\u00da",
               "tipo":"PPE",
               "valor":11602090.0
            },
            {
               "fundo":"BANCO ITA\u00da",
               "tipo":"CPR",
               "valor":5008949.0
            },
            {
               "fundo":"BANCO LUSO BRASILEIRO",
               "tipo":"CCE",
               "valor":8110010.0
            },
            {
               "fundo":"BANCO PINE",
               "tipo":"CCB",
               "valor":5468735.0
            },
            {
               "fundo":"BANCO SAFRA",
               "tipo":"CCE",
               "valor":405665.0
            },
            {
               "fundo":"BANCO SAFRA",
               "tipo":"CCE",
               "valor":700279.0
            },
            {
               "fundo":"BANCO SAFRA",
               "tipo":"CCE",
               "valor":1044083.0
            },
            {
               "fundo":"BANCO SAFRA",
               "tipo":"CCE",
               "valor":2013115.0
            },
            {
               "fundo":"BANCO SOFISA",
               "tipo":"CCB - Linha FGI",
               "valor":992064.0
            },
            {
               "fundo":"BANCO SOFISA",
               "tipo":"CCE",
               "valor":2464333.0
            },
            {
               "fundo":"UNIPRIME",
               "tipo":"CCB",
               "valor":1054789.0
            },
            {
               "fundo":"UNIPRIME",
               "tipo":"CCB",
               "valor":1704701.0
            }
         ],
         "request_id": "fa70b143-07e6-477d-bfd7-c6b2e18ea33a"
      },
	   {
         "nome":"sdaasd.pdf",
         "data":"dez/2024",
         "endividamento":[
            {
               "fundo":"UNIPRIME",
               "tipo":"CCE",
               "valor":1468451.0
            },
            {
               "fundo":"UNIPRIME",
               "tipo":"CCE",
               "valor":1084588.0
            },
            {
               "fundo":"BANCO C6 S.A.",
               "tipo":"CCB - Linha FGI",
               "valor":2148431.0
            },
            {
               "fundo":"BANCO C6 S.A.",
               "tipo":"CCE",
               "valor":3017354.0
            },
            {
               "fundo":"BANCO INDUSTRIAL",
               "tipo":"CCB - Linha FGI",
               "valor":1855771.0
            },
            {
               "fundo":"BANCO INDUSTRIAL",
               "tipo":"CCB - Linha FGI",
               "valor":369938.0
            },
            {
               "fundo":"CAIXA FEDERAL",
               "tipo":"CCB - Linha FGI",
               "valor":807745.0
            },
            {
               "fundo":"CAIXA FEDERAL",
               "tipo":"CCB - Linha FGI",
               "valor":402384.0
            },
            {
               "fundo":"CAIXA FEDERAL",
               "tipo":"CCB - Linha FGI",
               "valor":4045383.0
            },
            {
               "fundo":"SICOOB COCRED",
               "tipo":"CPRF",
               "valor":7054605.0
            },
            {
               "fundo":"SICOOB CREDMOTA",
               "tipo":"CCB",
               "valor":4458230.0
            },
            {
               "fundo":"SICOOB CREDMOTA",
               "tipo":"CPR",
               "valor":3165916.0
            },
            {
               "fundo":"BANCO STELLANTIS",
               "tipo":"CCB",
               "valor":399794.0
            },
            {
               "fundo":"BANCO STELLANTIS",
               "tipo":"CCB",
               "valor":412155.0
            },
            {
               "fundo":"BMW FINANCEIRA",
               "tipo":"CCB PRICE PARCELA BALAO",
               "valor":498773.0
            },
            {
               "fundo":"BANCO VOLKSWAGEN",
               "tipo":"CCB DAYCOVAL",
               "valor":282834.0
            },
            {
               "fundo":"BANCO SANTANDER",
               "tipo":"CPR",
               "valor":3767305.0
            },
            {
               "fundo":"BANCO SANTANDER",
               "tipo":"CPR",
               "valor":2907587.0
            },
            {
               "fundo":"BANCO SANTANDER",
               "tipo":"CCB",
               "valor":3085958.0
            },
            {
               "fundo":"QUAT\u00c1 INVESTIMENTOS",
               "tipo":"CCB",
               "valor":10057024.0
            },
            {
               "fundo":"LARCA CAPITAL",
               "tipo":"CESS\u00c3O DE CR\u00c9DITO",
               "valor":5000000.0
            },
            {
               "fundo":"BB CONSORCIO",
               "tipo":"FINAME",
               "valor":3589588.0
            },
            {
               "fundo":"BB CONSORCIO",
               "tipo":"FINAME",
               "valor":6816736.0
            },
            {
               "fundo":"BB CONSORCIO",
               "tipo":"FINAME",
               "valor":727429.0
            },
            {
               "fundo":"BB CONSORCIO",
               "tipo":"FINAME",
               "valor":5291060.0
            },
            {
               "fundo":"BB CONSORCIOS",
               "tipo":"CONSORCIO",
               "valor":161369.0
            }
         ],
          "request_id": "e3073a12-0136-454f-b87e-44ec4acdf60b"
      }
   ]
}