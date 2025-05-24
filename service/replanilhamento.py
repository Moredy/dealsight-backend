import httpx
from fastapi import UploadFile, File, Form, Depends, HTTPException
from fastapi.responses import Response
from sqlalchemy.orm import Session
from datetime import datetime
from dicttoxml import dicttoxml
from model.models import ReplanilhamentoHistorico, User
from typing import List, Optional
import logging
import re
import io
from datetime import datetime
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os
import json
logger = logging.getLogger(__name__)

MS_REPLANNING_URL = os.getenv('MS_REPLANNING_URL')

import io
import re
import httpx
from datetime import datetime
from fastapi import UploadFile, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
import logging

logger = logging.getLogger(__name__)


async def replanilhamento_db(
    files: List[UploadFile],
    user_id: int,
    cnpj_pagina: str,
    tipo: str,
    db,
    prompt: Optional[str] = None,
):
    try:
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            raise HTTPException(status_code=404, detail="Usuário não encontrado")


        files_payload = [
        ]

        for file in files:
            file_bytes = await file.read()
            files_payload.append(("files", (file.filename, file_bytes, file.content_type)))

        balanco_dict = {
            "Ativo Circulante": ["Caixa","Aplicações Financeiras","Contas a Receber","(-) PDD","Estoques","Adiant. A Fornecedores","Tributos a Recuperar","Despesas Antecipadas","Derivativos","Outros Créditos","Outros 1","Outros 2","Outros 3"],
            "Ativo Não Circulante": ["Partes Relacionadas","Tributos a Recuperar","Tributos Diferidos","Depósitos Judiciais","Aplicações Financeiras","Contas a Receber LP","Despesas Antecipadas","Ativo Mantido P/Venda","Derivativos","Outros Créditos","Outros 1","Outros 2","Outros 3"],
            "Ativo Permanente": ["Direito de Uso","Investimentos em Ligadas","Outros Investimentos","Ativos Biológicos","Imobilizado","(-) Depreciação","Intangível","(-) Amortização"],
            "Passivo Circulante": ["Bancos","Outras Dívidas Financeiras","Risco Sacado/Desc. Duplic.","Arrendamento","Fornecedores","Sal./Trib./Contrib.","Adiant. De Clientes","Provisão","Dividendos","Derivativos","Outros Débitos","Outros 1","Outros 2","Outros 3"],
            "Passivo Não Circulante": ["Bancos","Outras Dívidas Financeiras","Arrendamento","Fornecedores LP","Partes Relacionadas","Impostos Diferidos","Impostos Parcelados","Provisões","Dividendos","Derivativos","AFAC","Outros Créditos","Outros 1","Outros 2","Outros 3"],
            "Patrimônio Líquido": ["Capital Social","Reservas de Reavaliação","Reservas de Lucros","Outras Reservas","Lucros/Prejuízos","Acionistas Minoritários"]
        }

        payload_data = {
            "prompt": prompt or "",
            "balanco": json.dumps(balanco_dict, ensure_ascii=False)
        }

        async with httpx.AsyncClient(timeout=600.0) as client:
            response = await client.post(
                f"{MS_REPLANNING_URL}/{tipo}",
                files=files_payload,
                data=payload_data
            )

        if response.status_code != 200:
            raise HTTPException(status_code=response.status_code, detail=response.text)


        cnpj_pagina_limpo = re.sub(r'\D', '', cnpj_pagina)
        ms_replanning_request_id = response.json().get('request_id')


        # Salvar no banco
        novo_registro = ReplanilhamentoHistorico(
            ms_replanning_request_id=ms_replanning_request_id,
            cnpj=cnpj_pagina_limpo,
            responsavel=user.email,
            data_upload=datetime.utcnow(),
            tipo=tipo,
            organizacao_id=user.organizacao_id,
        )
        db.add(novo_registro)
        db.commit()

        # Retornar Excel gerado a partir do JSON
        return  {
        "id": novo_registro.id,
        "ms_replanning_request_id": novo_registro.ms_replanning_request_id,
        "cnpj": novo_registro.cnpj,
        "responsavel": novo_registro.responsavel,
        "data_upload": novo_registro.data_upload.isoformat(),
        "tipo": novo_registro.tipo,
        "organizacao_id": novo_registro.organizacao_id,
    }

    except Exception as e:
        logger.error(f"Erro no replanilhamento: {str(e)}")
        raise HTTPException(status_code=500, detail="Erro interno ao processar o replanilhamento.")

async def obter_historico_por_cnpj_db(cnpj: str,user_id: int,tipo: str, db):
    try:
        # Remove tudo que não for número
        cnpj_limpo = re.sub(r'\D', '', cnpj)

        # Busca o usuário e sua organização
        user = db.query(User).filter(User.id == user_id).first()
        if not user or not user.organizacao_id:
            raise HTTPException(status_code=403, detail="Usuário sem organização associada.")

        # Busca histórico filtrando pelo CNPJ e pela organização do usuário
        historico = (
            db.query(ReplanilhamentoHistorico)
            .filter(
                ReplanilhamentoHistorico.cnpj == cnpj_limpo,
                ReplanilhamentoHistorico.organizacao_id == user.organizacao_id,
                ReplanilhamentoHistorico.tipo == tipo
            )
            .order_by(ReplanilhamentoHistorico.data_upload.desc())
            .all()
        )



        historico_com_status = []

        for item in historico:
            data_replanning = await replanilhamento_status_db(item.ms_replanning_request_id, db)
            if data_replanning.get('status') == 'completed':
                historico_com_status.append({
                    "id": item.id,
                    "cnpj": item.cnpj,
                    "tipo": item.tipo,
                    "data_upload": item.data_upload,
                    "result": data_replanning.get("result"),
                })

        return historico_com_status

    except Exception as e:
        logger.error(f"Erro ao buscar histórico para o CNPJ {cnpj}: {str(e)}")
        raise HTTPException(status_code=500, detail="Erro interno ao buscar histórico da empresa.")


async def replanilhamento_status_db(request_id: str, db):
    """
    Consulta o status de replanilhamento via microsserviço externo.
    """
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            response = await client.get(f"{MS_REPLANNING_URL}/status/{request_id}")

        if response.status_code == 200:
            return response.json()

        raise HTTPException(status_code=response.status_code, detail=response.text)

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao consultar status: {str(e)}")

#
async def reclassificar_balanco_db(current_classification, prompt):
    try:

        balanco_dict = {
            "Ativo Circulante": ["Caixa","Aplicações Financeiras","Contas a Receber","(-) PDD","Estoques","Adiant. A Fornecedores","Tributos a Recuperar","Despesas Antecipadas","Derivativos","Outros Créditos","Outros 1","Outros 2","Outros 3"],
            "Ativo Não Circulante": ["Partes Relacionadas","Tributos a Recuperar","Tributos Diferidos","Depósitos Judiciais","Aplicações Financeiras","Contas a Receber LP","Despesas Antecipadas","Ativo Mantido P/Venda","Derivativos","Outros Créditos","Outros 1","Outros 2","Outros 3"],
            "Ativo Permanente": ["Direito de Uso","Investimentos em Ligadas","Outros Investimentos","Ativos Biológicos","Imobilizado","(-) Depreciação","Intangível","(-) Amortização"],
            "Passivo Circulante": ["Bancos","Outras Dívidas Financeiras","Risco Sacado/Desc. Duplic.","Arrendamento","Fornecedores","Sal./Trib./Contrib.","Adiant. De Clientes","Provisão","Dividendos","Derivativos","Outros Débitos","Outros 1","Outros 2","Outros 3"],
            "Passivo Não Circulante": ["Bancos","Outras Dívidas Financeiras","Arrendamento","Fornecedores LP","Partes Relacionadas","Impostos Diferidos","Impostos Parcelados","Provisões","Dividendos","Derivativos","AFAC","Outros Créditos","Outros 1","Outros 2","Outros 3"],
            "Patrimônio Líquido": ["Capital Social","Reservas de Reavaliação","Reservas de Lucros","Outras Reservas","Lucros/Prejuízos","Acionistas Minoritários"]
        }

        payload = {
            "prompt": prompt,
            "current_classification": current_classification,
            "balanco" : balanco_dict
        }

        async with httpx.AsyncClient(timeout=10.0) as client:
            response = await client.post(f"{MS_REPLANNING_URL}/balanco/reclassificar", json = payload)

        if response.status_code == 200:
            ms_replanning_request_id = response.json().get('request_id')
            return {"ms_replanning_request_id" : ms_replanning_request_id}

        raise HTTPException(status_code=response.status_code, detail=response.text)

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao consultar status: {str(e)}")

MESES_ORDENADOS = [
    "jan", "fev", "mar", "abr", "mai", "jun",
    "jul", "ago", "set", "out", "nov", "dez"
]

def extrair_mes_ano(mes_str):
    partes = mes_str.lower().split("/")
    if len(partes) == 2:
        return partes[0], int(partes[1])
    return None, None


def gerar_excel_faturamento(json_result: dict) -> StreamingResponse:
    cnpj_limpo = re.sub(r'\D', '', json_result.get('cnpj', 'cnpj_nao_informado'))
    dados = json_result.get("faturamento_mensal", [])

    faturamento_map = {}
    anos_set = set()
    for item in dados:
        mes_abrev, ano = extrair_mes_ano(item.get("mes", ""))
        if mes_abrev and ano:
            faturamento_map[(mes_abrev, ano)] = item.get("valor", 0)
            anos_set.add(ano)

    anos_ordenados = sorted(anos_set)

    wb = Workbook()
    ws = wb.active
    ws.title = "Faturamento Mensal"

    # Estilos
    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    month_fill = PatternFill(start_color="BF8F30", end_color="BF8F30", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center")

    # Cabeçalhos
    ws.cell(row=2, column=2, value="Mês").fill = header_fill
    ws.cell(row=2, column=2).font = white_font
    for idx, ano in enumerate(anos_ordenados):
        cell = ws.cell(row=2, column=3 + idx, value=str(ano))
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = center

    mes_valido_count = {}  # dict para armazenar quantos meses > 0 por ano

    for i, mes_abrev in enumerate(MESES_ORDENADOS):
        row = 3 + i
        ws.cell(row=row, column=2, value=mes_abrev.capitalize()).fill = month_fill
        ws.cell(row=row, column=2).font = bold_font
        for j, ano in enumerate(anos_ordenados):
            valor = faturamento_map.get((mes_abrev, ano), 0)
            col = 3 + j
            ws.cell(row=row, column=col, value=valor)
            if valor > 0:
                mes_valido_count[ano] = mes_valido_count.get(ano, 0) + 1

    total_row = 15
    media_row = 16

    ws.cell(row=total_row, column=2, value="Total").fill = month_fill
    ws.cell(row=total_row, column=2).font = bold_font
    ws.cell(row=media_row, column=2, value="Fat. Médio M").fill = month_fill
    ws.cell(row=media_row, column=2).font = bold_font

    for j, ano in enumerate(anos_ordenados):
        col = 3 + j
        total_cell = ws.cell(row=total_row, column=col)
        total_cell.value = f"=SUM({ws.cell(row=3, column=col).coordinate}:{ws.cell(row=14, column=col).coordinate})"
        total_cell.font = bold_font

        count = mes_valido_count.get(ano, 1)  # evita divisão por zero
        media_cell = ws.cell(row=media_row, column=col)
        media_cell.value = f"={total_cell.coordinate}/{count}"
        media_cell.font = bold_font

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"faturamento_{cnpj_limpo}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

def gerar_excel_endividamento(json_result: dict) -> StreamingResponse:
    import io
    import re
    from collections import defaultdict
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from fastapi.responses import StreamingResponse
    from fastapi import HTTPException
    import logging

    logger = logging.getLogger(__name__)

    MESES_MAP = {
        "jan": "Jan", "fev": "Feb", "mar": "Mar", "abr": "Apr",
        "mai": "May", "jun": "Jun", "jul": "Jul", "ago": "Aug",
        "set": "Sep", "out": "Oct", "nov": "Nov", "dez": "Dec"
    }

    def formatar_data(data_str: str):
        try:
            mes_pt, ano = data_str.strip().lower().split("/")
            mes_en = MESES_MAP.get(mes_pt)
            if mes_en:
                return f"{mes_en}-{ano[-2:]}"
        except:
            pass
        return "DataInválida"

    try:
        nome_arquivo = json_result.get("nome", "endividamento")
        data_ref = json_result.get("data", "")
        endividamento = json_result.get("endividamento", [])

        data_coluna = formatar_data(data_ref)

        wb = Workbook()

        # --- Aba 1: Endividamento detalhado ---
        ws_detalhado = wb.active
        ws_detalhado.title = "Endividamento"

        header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)
        center = Alignment(horizontal="center")

        safe_coluna = re.sub(r'[\\/*?:\[\]]', '-', data_coluna or "Período")

        headers = ["Banco", "Tipo", safe_coluna]
        for idx, col in enumerate(headers, start=1):
            cell = ws_detalhado.cell(row=1, column=idx, value=col)
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = center

        for i, item in enumerate(endividamento, start=2):
            banco = item.get("fundo", "").strip()
            tipo = item.get("tipo", "").strip()
            valor = float(item.get("valor", 0))

            ws_detalhado.cell(row=i, column=1, value=banco)
            ws_detalhado.cell(row=i, column=2, value=tipo)
            ws_detalhado.cell(row=i, column=3, value=round(valor, 2))

        # --- Aba 2: Agrupado por Banco e Tipo ---
        agrupado = defaultdict(float)
        for item in endividamento:
            banco = item.get("fundo", "").strip()
            tipo = item.get("tipo", "").strip()
            valor = float(item.get("valor", 0))
            agrupado[(banco, tipo)] += valor

        ws_agrupado = wb.create_sheet(title="Agrupado Banco-Tipo")

        for idx, col in enumerate(headers, start=1):
            cell = ws_agrupado.cell(row=1, column=idx, value=col)
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = center

        for i, ((banco, tipo), total) in enumerate(agrupado.items(), start=2):
            ws_agrupado.cell(row=i, column=1, value=banco)
            ws_agrupado.cell(row=i, column=2, value=tipo)
            ws_agrupado.cell(row=i, column=3, value=round(total, 2))

        # Gerar o arquivo final
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        nome_limpo = re.sub(r'\W+', '_', nome_arquivo.lower())
        filename = f"{nome_limpo}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        logger.error(f"Erro ao gerar Excel de endividamento: {str(e)}")
        raise HTTPException(status_code=500, detail="Erro ao gerar o Excel.")


def gerar_excel_balanco(json_result: dict) -> StreamingResponse:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    import io
    import re
    from collections import defaultdict

    cnpj_limpo = re.sub(r'\D', '', json_result.get('cnpj', 'cnpj_nao_informado'))
    balanco = json_result.get("balanco", {})

    wb = Workbook()
    ws = wb.active
    ws.title = "Balanço Patrimonial"

    # Estilos
    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    group_fill = PatternFill(start_color="BF8F30", end_color="BF8F30", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center")

    # Preparar estrutura
    valores = defaultdict(lambda: defaultdict(float))  # valores[tipo][data] = valor
    grupos = defaultdict(list)  # grupos[grupo] = lista de tipos

    datas_set = set()

    for grupo, registros in balanco.items():
        for reg in registros:
            tipo = reg.get("campo_novo")
            data = reg.get("mes")  # ex: "dez/2022"
            valor = reg.get("valor", 0)
            valores[tipo][data] += valor
            if tipo not in grupos[grupo]:
                grupos[grupo].append(tipo)
            datas_set.add(data)

    datas_ordenadas = sorted(datas_set, key=lambda d: (d[-4:], d[:3]))  # ordena por ano e mês

    # Cabeçalhos
    ws.cell(row=1, column=1, value="Ativo").fill = header_fill
    ws.cell(row=1, column=1).font = white_font

    for idx, data in enumerate(datas_ordenadas):
        col = 2 + idx
        ws.cell(row=1, column=col, value=data).fill = header_fill
        ws.cell(row=1, column=col).font = white_font
        ws.cell(row=1, column=col).alignment = center

    # Conteúdo
    row = 2
    for grupo, tipos in grupos.items():
        ws.cell(row=row, column=1, value=grupo).fill = group_fill
        ws.cell(row=row, column=1).font = bold_font
        row += 1

        for tipo in tipos:
            ws.cell(row=row, column=1, value=tipo)
            for idx, data in enumerate(datas_ordenadas):
                valor = valores[tipo].get(data, 0)
                ws.cell(row=row, column=2 + idx, value=valor)
            row += 1

    # Salvar e retornar
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"balanco_{cnpj_limpo}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def json_para_excel_db(json_result: dict, tipo: str) -> StreamingResponse:
    try:
        if tipo == "faturamento":
            return gerar_excel_faturamento(json_result)
        elif tipo == "endividamento":
            return gerar_excel_endividamento(json_result)
        elif tipo == "balanco":
            return gerar_excel_balanco(json_result)
        else:
            raise ValueError("Tipo de planilha inválido.")
    except Exception as e:
        logger.error(f"Erro ao gerar Excel ({tipo}): {str(e)}")
        raise HTTPException(status_code=500, detail="Erro ao gerar o Excel.")



